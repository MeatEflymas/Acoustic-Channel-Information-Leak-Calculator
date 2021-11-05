import os
import sys
import matplotlib as mpl
import matplotlib.pyplot as plt
import numpy as np
import openpyxl as xl
from openpyxl.drawing.image import Image
from openpyxl.styles import Alignment, Border, Side, PatternFill, Color
from openpyxl.utils.exceptions import InvalidFileException

from calc_acoustic_ch.data_handlers.data_loader import DataLoader
from calc_acoustic_ch.external_data.text_constant import *


class ReadersWriters:
    current_dir_abs_path = sys.path[0].replace('\\', '/')  # находим текущюю директорию
    if '/base_library.zip' in current_dir_abs_path:
        current_dir_abs_path = current_dir_abs_path[0:-17]

    IMAGES_FOLDER = current_dir_abs_path + '/calc_acoustic_ch/plot_images/'

    THIN_BORDER = Border(left=Side(style='thin'),
                         right=Side(style='thin'),
                         top=Side(style='thin'),
                         bottom=Side(style='thin'))

    CENTER_ALIGN = Alignment(horizontal='center',
                             wrap_text=True,
                             vertical='center')

    def __init__(self):
        pass

    # YES
    def generate_excel(self, path, count):
        workbook = xl.Workbook()
        workbook.remove_sheet(workbook.get_sheet_by_name('Sheet'))
        excel_cols = ['A', 'B', 'C', 'D', 'E', 'F']
        for index in range(0, count):
            iterator = COLUMN_NAMES_GEN.__iter__()
            sheet = workbook.create_sheet()
            for col in excel_cols:
                next_key = iterator.__next__()
                self._write_column_names(next_key, sheet, col, COLUMN_NAMES_GEN, 1.5)
            sheet.column_dimensions['A'].width = 10
            sheet['F1'].fill = PatternFill(start_color=Color(indexed=42), fill_type="solid")
        workbook.save(path)

    # YES
    # получение данных из excel файла с одного листа, представление их в виде numpy-массивов
    def read_excel_to_numpy_array_signal_noise(self, path, sheet_number=0):
        workbook = xl.load_workbook(path)
        sheet_names = workbook.get_sheet_names()
        name = sheet_names[sheet_number]
        sheet = workbook.get_sheet_by_name(name)
        eigenvalues_of_frequencies = False
        signal_level = []
        noise_level = []
        f_lower = []
        f_higher = []
        for row in sheet.iter_rows(min_row=2, min_col=2, max_col=3):
            if (row[0].value or row[1].value) is None:
                break
            signal_level.append(row[0].value)
            noise_level.append(row[1].value)
        if not sheet['D2'].value is None:
            eigenvalues_of_frequencies = True
            for row in sheet.iter_rows(min_row=2, min_col=4, max_col=5):
                if (row[0].value or row[1].value) is None:
                    break
                f_lower.append(row[0].value)
                f_higher.append(row[1].value)
        workbook.close()
        if eigenvalues_of_frequencies:
            return [name,
                    path,
                    np.array(signal_level),
                    np.array(noise_level),
                    True,
                    np.array(f_lower),
                    np.array(f_higher)]
        else:
            return [name,
                    path,
                    np.array(signal_level),
                    np.array(noise_level),
                    False]

    # тот же функционал, как у предыдущего метода, только для всех листов книги
    def read_all_sheets_with_signal_noise(self, path):
        workbook = xl.load_workbook(path)
        sheet_names = workbook.get_sheet_names()
        sheet_count = len(sheet_names)
        result_array = []
        for index in range(0, sheet_count):
            result_array.append(self.read_excel_to_numpy_array_signal_noise(path, index))
        workbook.close()
        return result_array

    # YES
    # запись результатов расчета в excel файл
    def write_results_to_excel(self, result, same_path=True, new_path=""):
        # выбираем файл записи
        if same_path:
            path = result.native_path
        else:
            path = new_path

        sheet_name = result.name

        # открываем книгу, если ее нет по данному адресу, создаем, либо сообщаем об ошибке
        try:
            workbook = xl.load_workbook(path)
        # неверное расширение
        except InvalidFileException:
            pass
        # если файла не существует, пробуем создать
        except FileNotFoundError:
            if os.path.isdir(os.path.split(path)[0]):
                workbook = xl.Workbook()
                sheet = workbook.create_sheet(sheet_name)
                workbook.remove_sheet(workbook.get_sheet_by_name('Sheet'))
            # файл создан
            else:
                # файл не удалось создать, неверный путь
                pass
        else:
            # файл найден
            try:
                # ищем лист
                sheet = workbook.get_sheet_by_name(sheet_name)
            except KeyError:
                # лист не найден, создаем
                sheet = workbook.create_sheet(sheet_name)

        # заполняем ячейки данными из массивов
        excel_cols_arrays = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I']
        iterator = COLUMN_NAMES.__iter__()
        for col in excel_cols_arrays:
            next_key = iterator.__next__()
            self._write_column_names(next_key, sheet, col, COLUMN_NAMES)
            for row in range(2, result.dimension + 2):
                sheet[col + str(row)].alignment = self.CENTER_ALIGN
                sheet[col + str(row)].border = self.THIN_BORDER
                sheet[col + str(row)] = result.data_to_write[next_key][row - 2]
        sheet.column_dimensions['A'].width = 10

        # заполняем ячейки скалярами (единичными значениями)
        excel_cols_scalar = ['J', 'K', 'L', 'M']
        for col in excel_cols_scalar:
            next_key = iterator.__next__()
            self._write_column_names(next_key, sheet, col, COLUMN_NAMES)
            sheet[col + '2'].alignment = self.CENTER_ALIGN
            sheet[col + '2'].border = self.THIN_BORDER
            sheet[col + '2'] = result.data_to_write[next_key]

        x = result.data_to_write['NUMBER']
        y1 = result.data_to_write['SIGNAL_LEVEL']
        y2 = result.data_to_write['NOISE_LEVEL']

        image_lin = self._plot_linear(x, y1, y2, sheet_name, False)

        row = result.dimension + 3
        sheet.add_image(Image(image_lin), 'A' + str(row))

        image_diag = self._plot_diagram(x, y1, y2, sheet_name, False)
        sheet.add_image(Image(image_diag), 'J' + str(row))

        image_delta = self._plot_diagram_delta(x, np.subtract(y1, y2), sheet_name, False)
        sheet.add_image(Image(image_delta), 'O1')

        sheet['J3'].border = self.THIN_BORDER
        sheet['J3'].alignment = self.CENTER_ALIGN
        loader = DataLoader()
        sheet['J3'] = loader.get_rec_from_yaml(result.data_to_write["W_S"], result.data_to_write["W_R"])
        sheet.merge_cells(start_row=3, start_column=10, end_row=9, end_column=13)
        workbook.save(path)

    # YES
    def write_all_results_to_excel(self, results, same_path=True, new_path=""):
        for result in results:
            self.write_results_to_excel(result, same_path, new_path)

    def _write_column_names(self, next_key, sheet, col, current_dict, divider=2.0):
        next_col_name = current_dict.get(next_key)
        sheet[col + '1'] = next_col_name
        sheet[col + '1'].alignment = self.CENTER_ALIGN
        sheet[col + '1'].border = self.THIN_BORDER
        sheet.column_dimensions[col].width = len(next_col_name) / divider

    def _plot_linear(self, x, y1, y2, name, with_title=True):
        plt.rcParams['figure.figsize'] = 8, 4
        if with_title:
            plt.title(PLOT_TITLE + name)
        plt.grid()
        plt.tick_params(axis='both',  # Применяем параметры к обеим осям
                        which='major',  # Применяем параметры к основным делениям
                        direction='inout',  # Рисуем деления внутри и снаружи графика
                        pad=10,  # Расстояние между черточкой и ее подписью
                        labelsize=10,  # Размер подписи
                        bottom=True,  # Рисуем метки снизу
                        top=True,  # сверху
                        left=True,  # слева
                        right=True,  # и справа
                        labelbottom=True,  # Рисуем подписи снизу  # сверху
                        labelleft=True)
        plt.plot(x, y1, label=PLOT_LABEL_1, linewidth=3)
        plt.plot(x, y2, label=PLOT_LABEL_2, linewidth=3)
        plt.xticks(x)
        plt.legend(loc='lower left')
        image = self.IMAGES_FOLDER + name + "_lin.png"
        plt.savefig(image)
        plt.close()
        return image

    def _plot_diagram(self, x, y1, y2, name, with_title=True):
        # mpl.rcParams.update({'font.size': 12})
        plt.rcParams['figure.figsize'] = 8, 4

        if with_title:
            plt.title(PLOT_TITLE + name)

        ax = plt.axes()
        ax.yaxis.grid(True, zorder=1)
        ax.xaxis.grid(True, zorder=1)

        xs = range(len(x))

        plt.bar([x - 0.20 for x in xs], y1,
                width=0.40, color='red', alpha=0.7, label=PLOT_LABEL_1,
                zorder=2)
        plt.bar([x + 0.20 for x in xs], y2,
                width=0.40, color='blue', alpha=0.7, label=PLOT_LABEL_2,
                zorder=2)
        plt.xticks(xs, x)

        plt.legend(loc='lower center')
        image = self.IMAGES_FOLDER + name + "_diag.png"
        plt.savefig(image)
        plt.close()
        return image

    def _plot_diagram_delta(self, x, y, name, with_title=True):
        # mpl.rcParams.update({'font.size': 12})
        plt.rcParams['figure.figsize'] = 8, 4

        if with_title:
            plt.title(PLOT_TITLE + name)

        ax = plt.axes()
        ax.yaxis.grid(True, zorder=1)
        ax.xaxis.grid(True, zorder=1)

        xs = range(len(x))

        plt.bar([x for x in xs], y,
                width=0.40, color='green', alpha=0.7, label=PLOT_LABEL_3,
                zorder=2)
        plt.xticks(xs, x)

        plt.legend(loc='lower left')
        image = self.IMAGES_FOLDER + name + "_diag_delta.png"
        plt.savefig(image)
        plt.close()
        return image
